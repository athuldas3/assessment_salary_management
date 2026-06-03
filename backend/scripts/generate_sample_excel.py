"""Generate the sample Excel file used for assessment reference."""

import sys
from pathlib import Path

BACKEND_ROOT = Path(__file__).resolve().parents[1]
PROJECT_ROOT = BACKEND_ROOT.parent
sys.path.insert(0, str(BACKEND_ROOT))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from app.services.seed_data import build_employee_record

OUTPUT_PATH = PROJECT_ROOT / "data" / "sample_employees.xlsx"
SAMPLE_COUNT = 50

HEADERS = [
    "Employee ID",
    "Full Name",
    "Country",
    "Job Title",
    "Department",
    "Annual Salary",
    "Notes",
]

NOTES = {
    0: "Sample export from legacy spreadsheet workflow",
    1: "Used for local demo/reference only",
    2: "Salary values are annual amounts in local currency units",
}


def generate_sample_excel() -> Path:
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Employees"

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for column, header in enumerate(HEADERS, start=1):
        cell = worksheet.cell(row=1, column=column, value=header)
        cell.fill = header_fill
        cell.font = header_font

    for index in range(SAMPLE_COUNT):
        record = build_employee_record(index)
        row = index + 2
        worksheet.cell(row=row, column=1, value=str(record.id))
        worksheet.cell(row=row, column=2, value=record.full_name)
        worksheet.cell(row=row, column=3, value=record.country)
        worksheet.cell(row=row, column=4, value=record.job_title)
        worksheet.cell(row=row, column=5, value=record.department)
        worksheet.cell(row=row, column=6, value=float(record.salary))
        worksheet.cell(row=row, column=7, value=NOTES.get(index, ""))

    for column in range(1, len(HEADERS) + 1):
        worksheet.column_dimensions[get_column_letter(column)].width = 24

    worksheet.freeze_panes = "A2"
    workbook.save(OUTPUT_PATH)
    return OUTPUT_PATH


if __name__ == "__main__":
    path = generate_sample_excel()
    print(f"Created {path} with {SAMPLE_COUNT} rows")
